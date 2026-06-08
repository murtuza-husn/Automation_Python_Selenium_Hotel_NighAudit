import msvcrt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime, timedelta, time
import re
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Path to save file on Desktop
desktop_path = r"C:\Users\econo\Downloads"
print(desktop_path)
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
url = "https://www.choiceadvantage.com/choicehotels/sign_in.jsp"
driver_path = r"C:\Program Files (x86)\chromedriver-win64\chromedriver.exe"
inHouseList_URL = "https://www.choiceadvantage.com/choicehotels/ViewInHouseList.init"
todaysCheckedOutGuest_URL = "https://www.choiceadvantage.com/choicehotels/ViewCheckedOutList.init"

def get_headless_driver():
    chrome_options = Options()
    #chrome_options.add_argument("--headless=new")  # Use 'new' headless mode
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1920x1080")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--new-window")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-infobars")

    try:
        # Automatically download the correct driver
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.get("https://www.google.com")
        print("Driver loaded successfully!")
        return driver
    except Exception as e:
        print("An error occurred:", e)
        return None

def get_password(prompt="Enter password: "):
    print(prompt, end="", flush=True)
    password = ""
    while True:
        ch = msvcrt.getch()
        if ch in {b'\r', b'\n'}:  # Enter key pressed
            print("")  # Move to next line
            break
        elif ch == b'\x08':  # Backspace
            if len(password) > 0:
                password = password[:-1]
                print("\b \b", end="", flush=True)
        else:
            password += ch.decode("utf-8")
            print("*", end="", flush=True)
    return password



def get_credentials(driver):

    userName = input("Enter Username : ")
    password = pwd = get_password()

    try:
        driver.get(url)
        print("After driver.get() URL is:", driver.current_url)
        print("Page title is:", driver.title)
        username_field = driver.find_element(By.NAME, "j_username")
        password_field = driver.find_element(By.NAME, "j_password")
        username_field.send_keys(userName)
        password_field.send_keys(password)
        login_button = driver.find_element(By.ID, "greenButton")
        login_button.click()
        WebDriverWait(driver, 25).until(
            EC.presence_of_element_located((By.ID, "infoNews"))
        )
        print("\nLogin Successful!")
        print("\n")
    except Exception as e:
        raise Exception("The Credentials entered are not valid. Please try again.") from e

def todaysCheckedOutGuest(driver):
    driver.get(todaysCheckedOutGuest_URL)
    checkout_count = 0
    tbody = driver.find_element("id", "checkedOutList")
    rows = tbody.find_elements("tag name", "tr")
    checkedout_room_numbers = []
    for row in rows:
        cells = row.find_elements("tag name", "td")
        if len(cells) >= 4:
            room_text = cells[3].text.strip()
            if room_text.isdigit():  # make sure it's a number
                checkedout_room_numbers.append(int(room_text))

    checkedout_room_numbers.sort()
    print("Today's Checked Out Guest List :")
    print("================================")
    for room in checkedout_room_numbers:
        checkout_count += 1
    print(f'The total No. of Checked Out guests : "{checkout_count}"')
    num = 1
    for room in checkedout_room_numbers:
        print(f"{num}) {room}")
        num += 1
    return checkedout_room_numbers

def todaysDate():
    now = datetime.now()
    current_time = now.time()

    # Define the window: 12:00 AM to 6:00 AM
    midnight = time(0, 0)
    six_am = time(8, 0)

    # Determine filter date
    if midnight <= current_time < six_am:
        filter_date = (now - timedelta(days=1)).date()
    else:
        filter_date = now.date()

    # Format filter_date string according to your table date format
    # Here I assume arrival dates are in "YYYY-MM-DD" format in the table cells
    filter_date_str = filter_date.strftime("%m/%d/%Y")
    print(f"==> Filtering Date : {filter_date_str}")
    print("\n")
    return filter_date_str

#Function for Check-In Rooms
def inHouseList(driver, filter_date_str):
    driver.get(inHouseList_URL)
    print("\n")
    # Locate the table body
    tbody = driver.find_element("id", "inHouseList")
    rows = tbody.find_elements("tag name", "tr")

    checkin_room_numbers = []
    arrival_count = 0
    # or better explicit wait here
    for row in rows:
        cells = row.find_elements("tag name", "td")
        if not cells:
            continue
        arrival_date_text = cells[9].text.strip()
        arrival_date_text = datetime.strptime(arrival_date_text, "%m/%d/%Y").strftime("%m/%d/%Y")
        try:
            if arrival_date_text == filter_date_str:
                room_text = cells[5].text.strip()
                if room_text.isdigit():
                    checkin_room_numbers.append(int(room_text))
        except Exception as e:
            print("No Matching entries found")
    # Sort the room numbers ascending
    checkin_room_numbers.sort()

    print(f"Today's Checked-In Guest List - Arrival Date : {filter_date_str}")
    print("=========================================================")
    for room in checkin_room_numbers:
        arrival_count += 1
    print(f'The Number of Arrivals for today {filter_date_str} is : "{arrival_count}"')
    num = 1
    for room in checkin_room_numbers:
        print(f"{num}) {room}")
        num += 1
    print("\n")
    return checkin_room_numbers

def checkedout_but_not_checkedin(checkin_room_numbers, checkedout_room_numbers):
    print("Rooms that got Checked-Out today, but didnt get Checked-In and are currently vacant :")
    print("=====================================================================================")
    checkedout_but_not_checkedin_list = []
    for room in checkedout_room_numbers:
        if room not in checkin_room_numbers:
            checkedout_but_not_checkedin_list.append(room)
    print(f'The Total no. of CheckedOut Rooms, that are not checked-in and are Vacant are : "{len(checkedout_but_not_checkedin_list)}"')
    print("Note : The following Rooms will be marked Vacant")
    num = 1
    for room in checkedout_but_not_checkedin_list:
        print(f"{num}) {room}")
        num += 1
    print("\n")

rooms_data = {}
def GuestTracking(driver, checkin_room_numbers):
    error_list = []
    driver.get(inHouseList_URL)
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "inHouseList"))
    )

    for in_room in checkin_room_numbers:
        print(f"This is the value of in_room that is selected : {in_room}")
        print(f"Processing room {in_room} : ")
        print("---------------------")

        try:
            # Refresh table each loop to avoid stale elements
            tbody = driver.find_element(By.ID, "inHouseList")
            rows = tbody.find_elements(By.TAG_NAME, "tr")

            room_found = False

            for row in rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                if not cells:
                    continue

                room_num = cells[5].text.strip()
                if room_num.isdigit() and int(room_num) == in_room:
                    room_found = True
                    # Default Values :
                    #name = "-"
                    #plan = "-"
                    #estimated_total_cost = "-"

                    # Open guest details
                    name = cells[1].text.strip()
                    # Clicking the "Name" Link to open Guest Records:
                    element_name = driver.find_element(By.LINK_TEXT, name)
                    element_name.click()
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "guestFolioEnabled"))
                    )
                    print(f"Guest Name is : {name}")

                    # Extract 'Rate plan'
                    plan = driver.find_element(By.ID, "ratePlan").text
                    plan = driver.find_element(By.ID, "ratePlan").text.strip()
                    exception_plans = ["SRD", "LCITY"]
                    print(f"Guest Plan: {plan}")

                    if plan in exception_plans:
                        if plan == "SRD":
                            estimated_total_cost = "-"
                            print(f"Estimated Total Cost: {estimated_total_cost}")

                            balance = "-"
                            print(f"Balance: {balance}")

                            # # Extract 'View Estimated Cost'
                            # # Get Inside Guest Folio - by clicking link
                            # driver.find_element(By.ID, "guestFolioEnabled").click()
                            # WebDriverWait(driver, 10).until(
                            #     EC.presence_of_element_located((By.ID, "button_12"))
                            # )
                            # # Get Inside View Estimated Cost - by clicking link
                            # driver.find_element(By.ID, "button_12").click()
                            # table = WebDriverWait(driver, 10).until(
                            #     EC.presence_of_element_located((By.XPATH, "//table"))
                            # )
                            # auth_rows = driver.find_elements(By.XPATH, "//table/tbody/tr")
                            #
                            # cards = []
                            # existing_auths = []
                            #
                            # # Extract Card and Existing Auth values for each row
                            # for authrow in auth_rows:
                            #     try:
                            #         card = authrow.find_element(By.XPATH, "./td[4]/p/em").text.strip()  # 4th td = Card
                            #         existing_auth = authrow.find_element(By.XPATH,
                            #                                              "./td[8]/p").text.strip()  # 8th td = Existing Auth
                            #         if existing_auth:  # only consider rows with values
                            #             existing_auth_value = float(existing_auth)
                            #             cards.append(card)
                            #             existing_auths.append(existing_auth_value)
                            #     except Exception as e:
                            #         # Skip rows without proper data
                            #         continue
                            #
                            # # Apply the 3-case logic
                            # selected_card = None
                            # selected_existing_auth = None
                            #
                            # if len(existing_auths) == 1:
                            #     # Case 1: only one row with a value
                            #     selected_card = cards[0]
                            #     selected_existing_auth = existing_auths[0]
                            # elif 25.00 in existing_auths:
                            #     # Case 2: one of the rows has 25.00
                            #     index = existing_auths.index(25.00)
                            #     selected_card = cards[index]
                            #     selected_existing_auth = existing_auths[index]
                            # else:
                            #     # Case 3: pick the row with the highest Existing Auth value
                            #     max_value = max(existing_auths)
                            #     index = existing_auths.index(max_value)
                            #     selected_card = cards[index]
                            #     selected_existing_auth = existing_auths[index]
                            #
                            # print(f"Selected Card: {selected_card}")
                            # print(f"Selected Existing Auth: {selected_existing_auth}")

                        if plan == "LCITY":
                            estimated_total_cost = "-"
                            balance = "-"
                            selected_card = "-"
                            selected_existing_auth = "-"
                            print(f"Estimated Total Cost: {estimated_total_cost}")
                            print(f"Balance: {balance}")
                            print(f"Selected Card: {selected_card}")
                            print(f"Selected Existing Auth: {selected_existing_auth}")


                    # If RatePlan is not SRD-Rate or City-of-Ottawa
                    else:
                        estimated_total_cost = "NULL"
                        estimated_total_cost = driver.find_element(By.ID, "estimated_total_cost").text
                        match = re.search(r"\d+(?:\.\d+)?", estimated_total_cost)
                        if match:
                            estimated_total_cost = match.group()
                            print(f"Estimated Total Cost: {estimated_total_cost}")
                        else:
                            print("Could not find estimated total cost:", estimated_total_cost)

                        # Extract 'Balance'
                        driver.find_element(By.ID, "guestFolioEnabled").click()
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.ID, "button_12"))
                        )
                        balance = driver.find_element(By.ID, "guestFolioBalance").text
                        match = re.search(r"\d+(?:\.\d+)?", balance)
                        balance = match.group()
                        if match:
                            print(f"Balance: {balance}")
                        else:
                            print("Could not find Balance:", balance)

                        # Extract 'View Estimated Cost'
                        driver.find_element(By.ID, "button_12").click()
                        table = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH, "//table"))
                        )
                        #Authorization and Card
                        auth_rows = driver.find_elements(By.XPATH, "//table/tbody/tr")

                        cards = []
                        existing_auths = []

                        # Extract Card and Existing Auth values for each row
                        for authrow in auth_rows:
                            try:
                                card = authrow.find_element(By.XPATH, "./td[4]/p/em").text.strip()  # 4th td = Card
                                existing_auth = authrow.find_element(By.XPATH,
                                                                 "./td[8]/p").text.strip()  # 8th td = Existing Auth
                                if existing_auth:  # only consider rows with values
                                    existing_auth_value = float(existing_auth)
                                    cards.append(card)
                                    existing_auths.append(existing_auth_value)
                            except Exception as e:
                                # Skip rows without proper data
                                continue

                        # Apply the 3-case logic
                        selected_card = None
                        selected_existing_auth = None

                        if len(existing_auths) == 1:
                            # Case 1: only one row with a value
                            selected_card = cards[0]
                            selected_existing_auth = existing_auths[0]
                        elif 25.00 in existing_auths:
                            # Case 2: one of the rows has 25.00
                            index = existing_auths.index(25.00)
                            selected_card = cards[index]
                            selected_existing_auth = existing_auths[index]
                        else:
                            # Case 3: pick the row with the highest Existing Auth value
                            max_value = max(existing_auths)
                            index = existing_auths.index(max_value)
                            selected_card = cards[index]
                            selected_existing_auth = existing_auths[index]

                        print(f"Selected Card: {selected_card}")
                        print(f"Selected Existing Auth: {selected_existing_auth}")

                    #wb = Workbook()
                    wb = load_workbook(file_path)
                    ws = wb.active

                    values = [in_room, name, plan, selected_card, selected_existing_auth, balance, "",
                               "", estimated_total_cost, "COVERED"]
                    ws.append(values)
                    # Save the workbook
                    wb.save(file_path)


                    # After finishing this room, go back to inHouseList page
                    driver.get(inHouseList_URL)
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "inHouseList"))
                    )
                    break  # Exit rows loop after finding the room
            print(f"Excel file saved at: {file_path}")
            print("\n")
            if not room_found:
                print(f"Room {in_room} not found in the current in-house list.")

        except Exception as e:
            error_list.append(in_room)
    return error_list

# def repeat(error_list, max_tries =3):
#     if len(error_list) != 0:
#         GuestTracking(driver, error_list)

file_name = "folio_list.xlsx"
file_path = os.path.join(desktop_path, file_name)

def workfile():
    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Room Payments"
    headers = ["Room No", "Name", "Plan", "Card Type", "Authorization", "Payment", "Deposit/Refund", "Balance", "Est. Total Cost", "Room Covered"]
    ws.append(headers)
    wb.save(file_path)

def retry_guest_tracking(driver, checkin_room_numbers, max_retries=20):
    """
    Repeatedly call GuestTracking on the remaining rooms until no errors or max_retries reached.
    """
    error_list = []
    remaining_rooms = checkin_room_numbers.copy()
    attempt = 1

    print("Printing Checkin Guest Details : ")
    print("================================")
    while remaining_rooms and attempt <= max_retries:
        print(f" -- Attempt #{attempt} for rooms: {remaining_rooms}\n")
        error_list = GuestTracking(driver, remaining_rooms)

        if not error_list:
            print("All rooms processed successfully!")
            break

        print(f"Rooms that failed in attempt #{attempt}: {error_list}")
        remaining_rooms = error_list
        attempt += 1

    if error_list:
        print(f"\nThe following rooms could not be processed after {max_retries} attempts: {error_list}")
    else:
        print(f"\nAll rooms processed successfully after {attempt} retries!")

def overwrite_workbook():
    pass

def main():
    driver = get_headless_driver()
    if driver is None:
        print("Failed to load driver, exiting.")
        return
    try:
        get_credentials(driver)
        todayDate = todaysDate()
        checkedout_rooms = todaysCheckedOutGuest(driver)
        checkin_rooms = inHouseList(driver, todayDate)
        checkedout_but_not_checkedin(checkin_rooms, checkedout_rooms)
        workfile()
        #vacant_rooms = vacant_list(checkin_rooms, checkedout_rooms)
        retry_guest_tracking(driver, checkin_rooms)

    finally:
        try:
            driver.quit()
        except Exception as e:
            print(f"Error quitting driver: {e}")


if __name__ == '__main__':
    main()
